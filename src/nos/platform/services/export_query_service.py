"""
Export Query Service - Export query results to CSV, Excel, JSON files.

Provides functionality for exporting query results to various formats and saving
to temporary files for download.

Usage:
    from nos.platform.services.export_query_service import export_query_service

    # Export query results
    result = export_query_service.export_query_result(
        columns=['id', 'name'],
        rows=[{'id': 1, 'name': 'Test'}],
        format='csv',
        filename_prefix='query_result'
    )
    # result.download_url -> '/api/download/temp/query_result_1234567890.csv'
"""

import os
import csv
import json
import logging
import time
from pathlib import Path
from typing import List, Dict, Any, Optional
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class ExportResult:
    """Result of an export operation."""
    success: bool
    file_path: Optional[str] = None
    download_url: Optional[str] = None
    filename: Optional[str] = None
    format: Optional[str] = None
    file_size: int = 0
    error: Optional[str] = None


class ExportQueryService:
    """
    Service for exporting query results to various file formats.

    Supported formats:
        - CSV: Comma-separated values
        - Excel: XLSX format (requires openpyxl)
        - JSON: Pretty-printed JSON
    """

    SUPPORTED_FORMATS = ['csv', 'excel', 'json']
    FORMAT_EXTENSIONS = {
        'csv': '.csv',
        'excel': '.xlsx',
        'json': '.json'
    }
    FORMAT_MIMETYPES = {
        'csv': 'text/csv',
        'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'json': 'application/json'
    }

    def __init__(self):
        self._temp_path: Optional[Path] = None
        self._initialized = False

    def _ensure_initialized(self):
        """Initialize the service on first use."""
        if self._initialized:
            return

        # Get temp path from environment
        temp_path_str = os.getenv("NOS_TEMP_PATH") or os.getenv("ORKESTRO_TEMP_PATH") or "{user_home}/.nos/temp"

        # Replace placeholders
        user_home = str(Path.home())
        temp_path_str = temp_path_str.replace('{user_home}', user_home)

        self._temp_path = Path(temp_path_str)

        # Ensure directory exists
        try:
            self._temp_path.mkdir(parents=True, exist_ok=True)
            logger.info(f"Export temp path: {self._temp_path}")
        except Exception as e:
            logger.error(f"Failed to create temp directory: {e}")
            # Fallback to system temp
            import tempfile
            self._temp_path = Path(tempfile.gettempdir()) / 'nos' / 'temp'
            self._temp_path.mkdir(parents=True, exist_ok=True)

        self._initialized = True

    @property
    def temp_path(self) -> Path:
        """Get the temporary files path."""
        self._ensure_initialized()
        return self._temp_path

    def generate_filename(
        self,
        prefix: str = 'export',
        format: str = 'csv',
        execution_id: Optional[str] = None,
        context_id: Optional[str] = None
    ) -> str:
        """
        Generate a unique filename for export.

        Args:
            prefix: Filename prefix
            format: Export format (csv, excel, json)
            execution_id: Optional execution ID
            context_id: Optional workflow/node ID

        Returns:
            Generated filename with extension
        """
        timestamp = int(time.time() * 1000)
        parts = [prefix]

        if execution_id:
            parts.append(execution_id[:8])
        if context_id:
            parts.append(context_id[:8])

        parts.append(str(timestamp))

        extension = self.FORMAT_EXTENSIONS.get(format.lower(), '.csv')
        return '_'.join(parts) + extension

    def export_query_result(
        self,
        columns: List[str],
        rows: List[Dict[str, Any]],
        format: str = 'csv',
        filename: Optional[str] = None,
        execution_id: Optional[str] = None,
        context_id: Optional[str] = None
    ) -> ExportResult:
        """
        Export query results to a file.

        Args:
            columns: List of column names
            rows: List of row dictionaries
            format: Export format (csv, excel, json)
            filename: Optional custom filename
            execution_id: Optional execution ID for filename
            context_id: Optional workflow/node ID for filename

        Returns:
            ExportResult with file path and download URL
        """
        self._ensure_initialized()

        format_lower = format.lower()
        if format_lower not in self.SUPPORTED_FORMATS:
            return ExportResult(
                success=False,
                error=f"Unsupported format: {format}. Supported: {', '.join(self.SUPPORTED_FORMATS)}"
            )

        # Generate filename if not provided
        if not filename:
            filename = self.generate_filename(
                prefix='query',
                format=format_lower,
                execution_id=execution_id,
                context_id=context_id
            )

        file_path = self.temp_path / filename

        try:
            if format_lower == 'csv':
                self._export_csv(file_path, columns, rows)
            elif format_lower == 'excel':
                self._export_excel(file_path, columns, rows)
            elif format_lower == 'json':
                self._export_json(file_path, columns, rows)

            file_size = file_path.stat().st_size
            download_url = f"/api/download/temp/{filename}"

            logger.info(f"Exported {len(rows)} rows to {file_path} ({file_size} bytes)")

            return ExportResult(
                success=True,
                file_path=str(file_path),
                download_url=download_url,
                filename=filename,
                format=format_lower,
                file_size=file_size
            )

        except Exception as e:
            logger.error(f"Export failed: {e}")
            return ExportResult(
                success=False,
                error=str(e)
            )

    def _export_csv(self, file_path: Path, columns: List[str], rows: List[Dict[str, Any]]):
        """Export to CSV format."""
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            for row in rows:
                writer.writerow([row.get(col, '') for col in columns])

    def _export_excel(self, file_path: Path, columns: List[str], rows: List[Dict[str, Any]]):
        """Export to Excel format (requires openpyxl)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row.get(col_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row in rows:
                cell_value = str(row.get(col_name, ''))
                max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        wb.save(file_path)

    def _export_json(self, file_path: Path, columns: List[str], rows: List[Dict[str, Any]]):
        """Export to JSON format."""
        export_data = {
            'columns': columns,
            'rows': rows,
            'row_count': len(rows),
            'exported_at': time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())
        }

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, default=str)

    def cleanup_old_files(self, max_age_hours: int = 24):
        """
        Remove old temporary files.

        Args:
            max_age_hours: Maximum age in hours before deletion
        """
        self._ensure_initialized()

        max_age_seconds = max_age_hours * 3600
        current_time = time.time()
        deleted_count = 0

        try:
            for file_path in self.temp_path.iterdir():
                if file_path.is_file():
                    file_age = current_time - file_path.stat().st_mtime
                    if file_age > max_age_seconds:
                        file_path.unlink()
                        deleted_count += 1

            if deleted_count > 0:
                logger.info(f"Cleaned up {deleted_count} old export files")

        except Exception as e:
            logger.error(f"Cleanup failed: {e}")

    def get_file_info(self, filename: str) -> Optional[Dict[str, Any]]:
        """
        Get information about an exported file.

        Args:
            filename: The filename to check

        Returns:
            Dictionary with file info or None if not found
        """
        self._ensure_initialized()

        file_path = self.temp_path / filename

        if not file_path.exists() or not file_path.is_file():
            return None

        stat = file_path.stat()
        extension = file_path.suffix.lower()

        # Determine format from extension
        format_map = {'.csv': 'csv', '.xlsx': 'excel', '.json': 'json'}
        format_type = format_map.get(extension, 'unknown')

        return {
            'filename': filename,
            'file_path': str(file_path),
            'format': format_type,
            'mimetype': self.FORMAT_MIMETYPES.get(format_type, 'application/octet-stream'),
            'size': stat.st_size,
            'created_at': stat.st_ctime,
            'modified_at': stat.st_mtime
        }


# Singleton instance
export_query_service = ExportQueryService()
